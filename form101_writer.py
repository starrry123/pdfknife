import openpyxl, io, os
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.pdfgen import canvas
from reportlab.lib.colors import black,blue,white,red,pink
from reportlab.lib.pagesizes import A4
#from reportlab.graphics import renderPDF

xls_text_coords=[[110, 525], [370, 525], [110, 490], [450, 490], [110, 470], [470, 470], [190, 380], [150, 160], [330, 160], [500, 160], [50, 373], [320, 373], [50, 350], [320, 350], [50, 325]]
sec3=[[120,435,'Woodside KGP'],[120,415,'Burrup Road'],[120,398,'Burrup'],[370,398,'WA'],[530,398,'6714']]
sec4=[[120,345,'Woodside'],[120,328,'Juan Nicklaus'],[370,328,'08 93486493'],[370,313,'Juan.Nicklaus@woodside.com.au'],[120,278,'Juan Nicklaus']]
sec5=[[120,243,'Karlak, 11 Mount Street'], [120,225,'Perth'],[370, 225,'WA'],[530,225,'6000']]
sec7=[[110,655,'Haitao Han'],[110,638,'94 Discovery Drive'],[110,620,'Bibra Lake'],[110,603,'0478848313'],[370,620,'WA'],[370,603,'haitao.han@applus.com'], [530,620,'6163']]
sec8=[[110,568,'Burrup Road'],[110,550,'Burrup'],[370,550,'WA'],[530,550,'6714'],[110,420,'Haitao Han']]

def GeneratePDF():
    xls=r'form101_list.xlsx'
    wb=openpyxl.load_workbook(filename=xls,read_only=False, keep_vba=True)
    ws=wb.worksheets[0]
    lastRow=4 # Default row 4 or ws.max_row
    while ws.cell(column=1, row=lastRow).value : lastRow+=1
    
    entry_no=lastRow-3 #
    col_num=16
    text=[['']]*col_num
 
    #read in coordinate to a list
    for i in range(entry_no):
        for j in range(col_num):
            cellvalue=ws.cell(row=3+i, column=j+1).value
            if cellvalue is not None:
                text[j]=cellvalue
            else:
                text[j]='UNKNOWN'
        asset_id=text[0]
    
        pdf_name='form101.pdf' # place the file under same path with this script  
        pdf_output=str(asset_id)+'_FORM101(DRAFT).pdf'
        write_pdf(pdf_name,pdf_output, text,xls_text_coords)
        
    wb.close()

def write_pdf(pdf_name,pdf_output,text,xls_text_coords):

    def fillout_static_text(canv,alist):
        for text_list in alist:
            x,y,text=text_list
            canv.drawString(x,y,text)

    def add_watermark(canv):
        canv.setFont("Helvetica-Bold", 16); canv.setFillColor(red)
        canv.setFillColor(pink)
        canv.setDash(3,2)
        canv.rect(395,670,190,100,stroke=1,fill=1)
        canv.setFillColor(red)
        canv.drawString(400,750,'DRAFT')
        canv.drawString(400,730,'DO NOT USE!')
        canv.drawString(400,690,str(text[0]))

    def mergepage(page_i,packet):
        packet.seek(0)
        pdf_buffer = PdfFileReader(packet)
        page=existing_pdf.getPage(page_i)
        page.mergePage(pdf_buffer.getPage(page_i))
        output.addPage(page)
        output.write(outputStream)


    #Page one addon text
    packet = io.BytesIO()
    canv = canvas.Canvas(packet, pagesize = A4)
    canv.setAuthor('Haitao Han')
    canv.setFont("Helvetica-Bold", 9); canv.setFillColor(blue);canv.setStrokeColor(blue)
    coords=[[0]*2]*14
    for i in range(0,10):
        coords=tuple(xls_text_coords[i])
        if i==2: #special case: join Asset ID and plant description
            canv.drawString(coords[0], coords[1]," ".join([text[0],text[3]]))
        elif i==6: # special case: append 'Room' to plant location
            canv.drawString(coords[0], coords[1], 'Room '+text[7])
        else:
            if text[i+1]=='UNKNOWN':
                canv.setFillColor(red)
            else:
                canv.setFillColor(blue)
            canv.drawString(coords[0], coords[1],str(text[i+1]))
    #Fill out common area text
    canv.setFont("Helvetica-Bold", 14)
    canv.drawString(135,645,'✓') # new registration
    if text[7] != 'UNKNOWN':
        canv.drawString(220,185,'✓') # mark as design rego
    else:
        canv.drawString(260,185,'✓') # mark as no design rego
    canv.setFont("Helvetica-Bold", 9)
    fillout_static_text(canv,sec3)
    fillout_static_text(canv,sec4)
    fillout_static_text(canv,sec5)
    add_watermark(canv)
    canv.showPage()

    #Page two addon text
    #canv = canv.canv(packet,pagesize = A4)
    canv.setFont("Helvetica-Bold", 9); canv.setFillColor(blue);canv.setStrokeColor(blue)
    coords=[[0]*2]*14
    pre_text=['Design Pressure: ', 'Design Temperature: ', 'Volume: ', 'Content Class: ', 'Hazard Level: ']
    for i in range(10,len(xls_text_coords)):
        coords=tuple(xls_text_coords[i])
        canv.drawString(coords[0], coords[1],pre_text[i-10]+text[i+1])
    fillout_static_text(canv,sec7)
    fillout_static_text(canv,sec8)
    add_watermark(canv)
    canv.save()
    
    outputStream = open(pdf_output, "wb")
    output = PdfFileWriter()
    output.addMetadata({
        '/Author': 'Haitao Han, haitao.han@applus.com',
        '/Title': 'WorkSafe Plant Registration Form 101',
        '/Subject':'Woodside KGP plant registration application form',
        '/Keywords': text[0],
        '/Producer': 'Hans PDF Generator'
    })
    if not pdf_name:
        outputStream.write(packet.getvalue())
    else:  
        existing_pdf = PdfFileReader(open(pdf_name, "rb"))
        for page_i in range(existing_pdf.numPages):
            mergepage(page_i,packet) 

    outputStream.close()
    #os.startfile(pdf_output,'open')

if __name__ == '__main__':
    GeneratePDF()
