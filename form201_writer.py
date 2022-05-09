# This script is used to bulk generate Western Australia Plant Registration form 201
import openpyxl, io, os,re
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.pdfgen import canvas
from reportlab.lib.colors import black,blue,white,red, Color
from reportlab.lib.pagesizes import A4

xls_text_coords=[[130,450],[330,450],[130,423],[460,423],[130,400],[460,400],[130,335],
    [130,290],[300,290],[460,290],
    [50,242],[300,242],[50,218],[300,218],[50,195],
    [130,135]]
sec2=[[130,380,'BURRUP ROAD'],[130,360,'BURRUP'],[330,360,'WA'],[460,360,'6714']]
sec678=[[440,652,'WOODSIDE ENERGY LTD.'],[440,635,'JUAN NICKLAUS'],[440,615,'WOODSIDE ENERGY LTD.'],[440,595,'005 482 986'],
                [330,572,'11 MOUNT ST'],[330,552,'PERTH'],[510,552,'6000'],[330,530,'JUAN.NICKLAUS@WOODSIDE.COM.AU'],[330,512,'08 9348 4000'],
                [313,405,'X'],[60,325,'JUAN.NICKLAUS@WOODSIDE.COM.AU'], [60,198,'JUAN NICKLAUS']]
sec_payment=[[215,585],
[95, 558], [115, 558], [135, 558], [155, 558],
[195, 558], [215, 558], [235, 558], [255, 558],
[295, 558], [315, 558], [335, 558], [355, 558],
[395, 558], [415, 558], [435, 558], [452, 558],
[100,533],
[95,500],[115,500], [158,500],[175,500],
[210,445]]
payment_text=['X',
    '4','8','6','5','0','1','1','0','0','0','4','7','1','2','0','4',
    'SHAYE STYLE',
    '0','2','2','3',
    '08 9348 4000']
payment_info_list=[]
for i in range(len(sec_payment)):
    x,y=sec_payment[i]
    payment_info_list.append([x,y,payment_text[i]])

def GeneratePDF():
    xls=r'form_list.xlsx'
    wb=openpyxl.load_workbook(filename=xls,read_only=True, keep_vba=True)
    ws=wb.worksheets[0]
    lastRow=4 # Default row 4 or ws.max_row
    while ws.cell(column=1, row=lastRow).value : lastRow+=1
    
    entry_no=lastRow-3 #
    col_num=17
    text=[['']]*col_num
 
    #read in coordinate to a list
    for i in range(entry_no):
        register_flag=ws.cell(row=3+i,column=18).value
        if register_flag is not None:
            print(ws.cell(row=3+i,column=1).value + '  SKIPPING...' +register_flag)
        else:
            for j in range(col_num):
                cellvalue=ws.cell(row=3+i, column=j+1).value
                if cellvalue is not None:
                    text[j]=cellvalue
                else:
                    text[j]='UNKNOWN'
            asset_id=text[0]
        
            pdf_name='form201.pdf' # place the file under same path with this script  
            pdf_output=re.sub('^AU01.','',asset_id)+'_FORM201(DRAFT).pdf'
            print("Generating File: "+pdf_output+' '+asset_id+' '+plant_location(asset_id))
            write_pdf(pdf_name,pdf_output, text,xls_text_coords)
            
    wb.close()

def plant_location(asset_id):
    pat=re.match(r'AU01\.(\d*)([a-zA-Z]+)(\d+)',asset_id)
    loc='Karratha Gas Plant '
    if pat is not None:
        if len(pat.group(1))==3:
            unit=pat.group(1)[1:3]
        else:
            unit=pat.group(3)[0:2]
        if pat.group(1).isnumeric():
            loc+=' Train '+pat.group(1)+' Unit '+unit+'00'
        elif pat.group(2) in ['A','GT']:
            loc+='Utility Unit '+ unit +'00'
        else:
            loc+='Unit '+ unit +'00'
        return loc

def write_pdf(pdf_name,pdf_output,text,xls_text_coords):

    def fillout_static_text(canv,alist):
        for text_list in alist:
            x,y,text=text_list
            canv.drawString(x,y,text)

    def add_watermark(canv):
        canv.setFont("Helvetica-Bold", 14)
        bg_transparent=Color(100,0,0,alpha=0.4)
        canv.setFillColor(red)
        canv.setFillColor(bg_transparent)
        canv.setDash(3,2)
        canv.rect(400,740,190,65,stroke=0,fill=1)
        canv.setFillColor(red)
        canv.drawString(410,785,'DRAFT -- DO NOT USE')
        canv.drawString(410,750,str(text[0]))

    def mergepage(page_i,packet):
        packet.seek(0)
        pdf_buffer = PdfFileReader(packet)
        page=existing_pdf.getPage(page_i)
        page.mergePage(pdf_buffer.getPage(page_i))
        output.addPage(page)
        output.write(outputStream)

    #Add 1st page text
    pre_text=['Design Pressure: ', 'Design Temperature: ', 'Volume: ', 'Content Class: ', 'Hazard Level: ']
    packet = io.BytesIO()
    canv = canvas.Canvas(packet, pagesize = A4)
    canv.setAuthor('Haitao Han')
    canv.setFont("Helvetica-Bold", 9); canv.setStrokeColor(blue)
    coords=[[0]*2]*14
    for i in range(0,len(xls_text_coords)):
        if text[i+1]=='UNKNOWN':
            canv.setFillColor(red)
        else:
            canv.setFillColor(blue)
        coords=tuple(xls_text_coords[i])
        if i <10:
            if i==2: #special case: join Asset ID and plant description
                canv.drawString(coords[0], coords[1]," ".join([text[0],text[3]]))
            elif i==6: # special case: append 'Room' to plant location
                canv.drawString(coords[0], coords[1], plant_location(text[0]))
            else:

                canv.drawString(coords[0], coords[1],str(text[i+1]))
        elif i==15:
            canv.drawString(coords[0], coords[1],str(text[i+1]))
        else:
            canv.drawString(coords[0], coords[1],pre_text[i-10]+text[i+1])
    canv.setFillColor(blue)
    canv.setFont("Helvetica-Bold", 14);canv.drawString(30,660,'X')
    canv.setFont("Helvetica-Bold", 9);canv.drawString(72,580,'X')
    fillout_static_text(canv,sec2)
    add_watermark(canv)
    canv.showPage()

    #Add 2nd page text: Applicant details
    canv.setFont("Helvetica-Bold", 9); canv.setFillColor(blue);canv.setStrokeColor(blue)
    canv.drawString(310,675,'X') #mark body corporate/company
    fillout_static_text(canv,sec678)
    add_watermark(canv)
    canv.showPage()
    
    #Add 3rd page text: Payment details
    canv.setFont("Helvetica-Bold", 14); canv.setFillColor(blue)
    fillout_static_text(canv,payment_info_list)
    canv.showPage()
    canv.save()

    outputStream = open(pdf_output, "wb")
    output = PdfFileWriter()
    output.addMetadata({
        '/Author': 'Haitao Han, haitao.han@applus.com',
        '/Title': 'WorkSafe Plant Registration Form 201',
        '/Subject':'Woodside KGP plant registration application form',
        '/Keywords': text[0]
    })
    if not pdf_name:
        outputStream.write(packet.getvalue())
    else:  
        existing_pdf = PdfFileReader(open(pdf_name, "rb"))
        for page_i in range(existing_pdf.numPages):
            mergepage(page_i,packet) 

    outputStream.close()
    os.startfile(pdf_output,'open')

if __name__ == '__main__':
    GeneratePDF()
