import openpyxl, io, os
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.pdfgen import canvas
from reportlab.lib.colors import black,blue,white,red,pink
from reportlab.lib.pagesizes import A4
#from reportlab.graphics import renderPDF

xls_pos=[[110, 525], [370, 525], [110, 490], [450, 490], [110, 470], [470, 470], [190, 380], [150, 160], [330, 160], [500, 160], [50, 373], [320, 373], [50, 350], [320, 350], [50, 325]]
sec3=[[120,435,'Woodside KGP'],[120,415,'Burrup Road'],[120,398,'Burrup'],[370,398,'WA'],[530,398,'6714']]
sec4=[[120,345,'Woodside'],[120,328,'Juan Nicklaus'],[370,328,'08 93486493'],[370,313,'Juan.Nicklaus@woodside.com.au'],[120,278,'Juan Nicklaus']]
sec5=[[120,243,'Karlak, 11 Mount Street'], [120,225,'Perth'],[370, 225,'WA'],[530,225,'6000']]
sec7=[[110,655,'Haitao Han'],[110,638,'94 Discovery Drive'],[110,620,'Bibra Lake'],[110,603,'0478848313'],[370,620,'WA'],[370,603,'haitao.han@applus.com'], [530,620,'6163']]
sec8=[[110,568,'Burrup Road'],[110,550,'Burrup'],[370,550,'WA'],[530,550,'6714'],[110,420,'Haitao Han']]

def get_xls(XLS):
    wb=openpyxl.load_workbook(filename=XLS,read_only=False, keep_vba=True)
    ws=wb.worksheets[0]
    lastRow=4 # Default row 4 or ws.max_row
    while ws.cell(column=1, row=lastRow).value :
        if ws.cell(column=1, row=lastRow).value != '':
            last_WIP_row=lastRow
        lastRow+=1
    entry_no=lastRow-3
    col_num=16
    text=[['']]*col_num
    pdf_name=r'form101.pdf'    
    #read in coordinate to a list

    for i in range(entry_no):
        for j in range(col_num):
            cellvalue=ws.cell(row=3+i, column=j+1).value
            if cellvalue is not None:
                text[j]=cellvalue
            else:
                text[j]='EMPTY_TBA'
        asset_id=text[0]

        pdf_output=str(asset_id)+'_FORM101(DRAFT).pdf'
        write_pdf(pdf_name,pdf_output, text,xls_pos)
        
    wb.close()

def write_pdf(pdf_name,pdf_output,text,xls_pos):

    def fillout(c,alist):
        for i in alist:
            x,y,text=i[0],i[1],i[2]
            c.drawString(x,y,text)
    def add_watermark(c):
        c.setFont("Helvetica-Bold", 16); c.setFillColor(red)
        c.setFillColor(pink)
        c.setDash(3,2)
        c.rect(395,670,190,100,stroke=1,fill=1)
        c.setFillColor(red)
        c.drawString(400,750,'DRAFT')
        c.drawString(400,730,'DO NOT USE!')
        c.drawString(400,690,str(text[0]))


    #Page one addon text
    packet1 = io.BytesIO()
    c1 = canvas.Canvas(packet1,pagesize = A4)
    c1.setAuthor('Haitao Han')
    c1.setFont("Helvetica-Bold", 9); c1.setFillColor(blue);c1.setStrokeColor(blue)
    coords=[[0]*2]*14
    for i in range(0,10):
        coords=tuple(xls_pos[i])
        if i==2: #special case: join Asset ID and plant description
            c1.drawString(coords[0], coords[1]," ".join([text[0],text[3]]))
        elif i==6: # special case: append 'Room' to plant location
            c1.drawString(coords[0], coords[1], 'Room '+text[7])
        else:
            if text[i+1]=='EMPTY_TBA':
                c1.setFillColor(red)
            else:
                c1.setFillColor(blue)
            c1.drawString(coords[0], coords[1],str(text[i+1]))
    #Fill out common area text
    c1.setFont("Helvetica-Bold", 14)
    c1.drawString(135,645,'✓') # new registration
    if text[7] != 'EMPTY_TBA':
        c1.drawString(220,185,'✓') # mark as design rego
    else:
        c1.drawString(260,185,'✓') # mark as no design rego
    c1.setFont("Helvetica-Bold", 9)
    fillout(c1,sec3)
    fillout(c1,sec4)
    fillout(c1,sec5)
    add_watermark(c1)
    c1.save()

    #Page two addon text
    packet2 = io.BytesIO()
    c2 = canvas.Canvas(packet2,pagesize = A4)
    c2.setFont("Helvetica-Bold", 9); c2.setFillColor(blue);c2.setStrokeColor(blue)
    coords=[[0]*2]*14
    pre_text=['Design Pressure: ', 'Design Temperature: ', 'Volume: ', 'Content Class: ', 'Hazard Level: ']
    for i in range(10,len(xls_pos)):
        coords=tuple(xls_pos[i])
        c2.drawString(coords[0], coords[1],pre_text[i-10]+text[i+1])
    fillout(c2,sec7)
    fillout(c2,sec8)
    add_watermark(c2)

    c2.save()
    
    packet1.seek(0)
    pdf_buffer = PdfFileReader(packet1)
    outputStream = open(pdf_output, "wb")
    output = PdfFileWriter()
    output.addMetadata({
        '/Author': 'Haitao Han, haitao.han@applus.com',
        '/Title': 'WorkSafe Plant Registration Form 101',
        '/Subject':'Woodside KGP plant registration application form',
        '/Keywords': text[0],
        '/Producer': 'Haitao PDF Generator'

    })
    if not pdf_name:
        outputStream.write(packet.getvalue())
    else:  
        existing_pdf = PdfFileReader(open(pdf_name, "rb"))
        page = existing_pdf.getPage(0)
        page.mergePage(pdf_buffer.getPage(0))
        output.addPage(page)
        output.write(outputStream)
        packet2.seek(0)
        pdf_buffer2 = PdfFileReader(packet2)
        page2=existing_pdf.getPage(1)
        page2.mergePage(pdf_buffer2.getPage(0))
        output.addPage(page2)
        output.write(outputStream)
       
    outputStream.close()
    #os.startfile(pdf_output,'open')

if __name__ == '__main__':
    pdf_name=r'form_101_0.pdf'
    xls=r'form101_list.xlsx'
    get_xls(xls)

