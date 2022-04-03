# This script is used to calculate MHF hazard level
import openpyxl, re
from prettytable import PrettyTable

def hazard_level_check():
    xls=r'form101_list.xlsx'
    wb=openpyxl.load_workbook(filename=xls,read_only=False, keep_vba=True)
    ws=wb.worksheets[0]
    lastRow=4 # Default row 4 or ws.max_row
    while ws.cell(column=1, row=lastRow).value : lastRow+=1
    
    entry_no=lastRow-3 #
    col_num=16
    text=[['']]*col_num
    ptb=PrettyTable()
    ptb.field_names=['No','ID','DP','Vol','Content','CAL FORMULA','Site HZL','Cal HZL','HZL Match?']
 
    for i in range(entry_no):
        for j in range(col_num):
            cellvalue=ws.cell(row=3+i, column=j+1).value
            if cellvalue is not None:
                text[j]=cellvalue
            else:
                text[j]='UNKNOWN'
        id=text[0]
        dp=text[11]
        vol=text[13]
        content=text[14]
        hzl=text[15]
        chamber_no_p=len(re.findall('(?i)\d+\s*[km]pa',dp))
        chamber_no_v=len(re.findall('(?i)\d+\s*L',vol))
        if chamber_no_p !=1 or chamber_no_v!=1:
            hzl_cal='--ERROR2--'
            match_flag='--ERROR2--'
            formula_text='--ERROR2--'

        else:
            dp_re=re.search('(\d+)\s*\w+',dp)
            if dp_re is not None:
                dp_num=float(dp_re.group(1))/1000 #Design pressure in MPa
            else:
                dp_num=0
            vol_re=re.search('(\d+)\s*L',vol)
            if vol_re is not None:
                vol_num=float(vol_re.group(1))
            else:
                vol_num=0
        
            pv=dp_num*vol_num

            Ff=1 #Fluid factor
            if content.startswith('H'):
                Ff=3
            elif content.startswith('VH'):
                Ff=10
            elif content.startswith('L'):
                Ff=1000    
            elif content=='NHL':
                Ff=round(1/3,3)
            elif content=='NHG':
                Ff=1
            Fc=1
            if content.endswith('G'):
                Fc=10
            Fs=3 #location: Major hazard facility
            if dp_num>50:
                Fs*=30

            h=pv*Fc*Ff*Fs
            if h>10**8.5:
                hzl_cal='A'
            elif h>=10**4:
                hzl_cal='B'
            elif h>=1000:
                hzl_cal='C'
            elif h>=10**2.5:
                hzl_cal='D'
            elif h==0:
                hzl_cal='--ERROR1--'
            else:
                hzl_cal='E'
            if hzl_cal in hzl:
                match_flag='Yes'
            elif dp_num==0 or vol_num==0:
                match_flag='--ERROR1--'
            else:
                match_flag='No'
            formula_text="*".join([str(dp_num),str(vol_num),str(Fc),str(Ff),str(Fs)])+'='+str(round(dp_num*vol_num*Fc*Ff*Fs,1))
        ptb.add_row([i+1,id,dp,vol,content,formula_text,hzl,hzl_cal,match_flag])

    print(ptb)
    wb.close()

if __name__ == '__main__':
    hazard_level_check()
